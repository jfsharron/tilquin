"""
================================================================================
 Program:           tilquin_accounts.py
 Software Engineer: Jonas Sharron
 Date:              19-April-2025

 Purpose:   Import existing Quicken accounts into tilquin database
================================================================================
"""
import sys
import re
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import mysql.connector
from mysql.connector import Error
import functools
import xlsxwriter
import pandas.io.sql as sql
import numpy as np
from numpy import loadtxt
from prettytable import from_db_cursor
from prettytable import PrettyTable
import datetime
from datetime import date
import os
from termcolor import colored, cprint 
from colorama import Fore, Back, Style 
from tabulate import tabulate
import fpdf
import colorama
from colorama import Fore, Back, Style
import getopt
from reportlab.lib.pagesizes import LETTER
from reportlab.pdfgen.canvas import Canvas
from reportlab.lib.units import inch
import webbrowser
import time
from pretty_html_table import build_table
from sqlalchemy import create_engine
import inquirer
from copy import copy
import re
from itertools import islice

# ==============================================================================
# intialize lists and variables
# ============================================================================== 

DATAFILE    = '/data/share/tilquin/accounts.xlsx'  
DF          = '/data/share/tilquin/account_dframe.xlsx'


# ==============================================================================
# establish database connection
# ==============================================================================

XUSER       = 'jfsharron'
XWORD       = 'marie151414'
HOST        = '192.168.2.107'
DATABASE    = 'tilquin'

try:
    CONNECTION = mysql.connector.connect(user=XUSER, password=XWORD,
    host=HOST, database=DATABASE)
    if CONNECTION.is_connected():
        db_Info = CONNECTION.get_server_info()
        print("Connected to MySQL Server version ", db_Info)
        global cursor
        cursor = CONNECTION.cursor()
        cursor.execute("select database();")
        record = cursor.fetchone()
        print("You're connected to database: ", record)
except Error as e:
    print("Error while connecting to MySQL", e)

# ==============================================================================
# import existing data from Quicken report
# ==============================================================================

# Load Excel
df = pd.read_excel(DATAFILE)

# Connect to MySQL
#conn = mysql.connector.connect(
#    host="localhost",
#    user="your_username",
#    password="your_password",
#    database="your_database"
#)

cursor = CONNECTION.cursor()

# Optional: Create table dynamically (depends on structure)
# Insert data
for _, row in df.iterrows():
    cursor.execute(
        "INSERT INTO account (account_name) VALUES (%s)",
        tuple(row)
    )

CONNECTION.commit()
cursor.close()
CONNECTION.close()






#column_names = ["Account"]
#data = pd.read_excel(DATAFILE, sheet_name = 'Report')
#df = pd.DataFrame(data)

#engine = create_engine('mysql+pymysql://jfsharron:marie151414@192.168.2.107/tilquin')




#wb = openpyxl.load_workbook(DATAFILE)
#sheet = wb['Report']
#
#column_names = ["Account"]
#data = pd.read_excel(DATAFILE, sheet_name = 'Report')
#df = pd.DataFrame(data)
#last_row_index =(len(df) + 2)
#for index, row in islice(df.iterrows(), 1, (last_row_index)): 
#    #print(row)
#    account = sheet.cell(row = index, column = 1).value
#    #account = ((row['Account']))
#    print(account)
#    print(last_row_index)
#    #account = str(account)








    #acct = (account)
    #save_query = ("INSERT INTO account (account_name)"
    #             "VALUES (%s)")
    #cursor = CONNECTION.cursor()
    #cursor.execute(save_query, account)
    #CONNECTION.commit()          