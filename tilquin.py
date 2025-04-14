"""
================================================================================
 Program:           tilquin.py
 Software Engineer: Jonas Sharron
 Date:              31-March-2025

 Purpose:   Populate database with data exported into Excel spreadsheet from
            Quicken for analysis, reporting, and functional usage.
================================================================================
"""
import sys
import re
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import mysql.connector
from mysql.connector import Error
import functools
import xlsxwriter
import pandas.io.sql as sql
import numpy as np
from numpy import loadtxt
from prettytable import from_db_cursor
from prettytable import PrettyTable
import datetime
from datetime import date
import os
from termcolor import colored, cprint 
from colorama import Fore, Back, Style 
from tabulate import tabulate
import fpdf
import colorama
from colorama import Fore, Back, Style
import getopt
from reportlab.lib.pagesizes import LETTER
from reportlab.pdfgen.canvas import Canvas
from reportlab.lib.units import inch
import webbrowser
import time
from pretty_html_table import build_table
from sqlalchemy import create_engine
import inquirer
from copy import copy
import re
from itertools import islice



# ==============================================================================
# intialize lists and variables
# ============================================================================== 

DATAFILE    = '/data/share/tilquin/atad2.xlsx'  
DF          = '/data/share/tilquin/dframe.xlsx'


# ==============================================================================
# establish database connection
# ==============================================================================

XUSER       = 'jfsharron'
XWORD       = 'marie151414'
HOST        = '192.168.2.107'
DATABASE    = 'tilquin'

try:
    CONNECTION = mysql.connector.connect(user=XUSER, password=XWORD,
    host=HOST, database=DATABASE)
    if CONNECTION.is_connected():
        db_Info = CONNECTION.get_server_info()
        print("Connected to MySQL Server version ", db_Info)
        global cursor
        cursor = CONNECTION.cursor()
        cursor.execute("select database();")
        record = cursor.fetchone()
        print("You're connected to database: ", record)
except Error as e:
    print("Error while connecting to MySQL", e)


def preProcess2():
    """
    ============================================================================
    Function:       preProcess2()
    Purpose:        format DATAFILE and import into database
    Parameter(s):   -None- 
    Return:         -None- 
    ============================================================================
    """ 

    """
    ============================================================================
    This portion of code is reponsible for formatting the input spreadsheet 
    correctly for import into database
    ============================================================================
    """
    # set intial values and define files
    # ================================================    
    column_names = ["InternalNo", "Date", "Account", "Num", "Description", 
                    "Memo", "Category", "Tag", "Notes", "Clr", "Amount"]
    intNo = 1001
    data = pd.read_excel(DATAFILE, sheet_name = 'Report')
    df0 = pd.DataFrame(data)

    # define rows to be used
    # =======================
    df1 = df0.drop ([1, 2, 3, 4, 5])
    last_row_index = (len(df1))
    beginLast_row_index = (last_row_index - 7)
    df = df1.drop(df1.index[beginLast_row_index : last_row_index]) 

    df.columns = column_names
    
    df['Date'] = pd.to_datetime(df['Date']).dt.date

    df.to_excel('dataframe3.xlsx', index = False)

    wb = openpyxl.load_workbook('dataframe3.xlsx')
    sheet = wb['Sheet1']

    # iterate through input file setting absent data values
    # ======================================================
    for index, row in islice(df.iterrows(), 1, (beginLast_row_index)): 

        # import aguements for data with date (no split or first row of split)
        # =====================================================================
        if not pd.isna(row['Date']):
            intNo                                   = intNo + 1
            ind                                     = (index - 3)
            sheet.cell(row = ind, column = 1).value = intNo 
            lastInt                                 = intNo
            lastDate                                = ((row['Date']))
            lastAcct                                = ((row['Account']))
            lastNum                                 = ((row['Num']))
            lastDesc                                = ((row['Description']))
            lastNotes                               = ((row['Notes']))

        elif pd.isna(row['Date']):
            ind = (index - 3)  
            sheet.cell(row = ind, column = 1).value = lastInt
            sheet.cell(row = ind, column = 2).value = lastDate
            sheet.cell(row = ind, column = 3).value = lastAcct
            sheet.cell(row = ind, column = 4).value = lastNum
            sheet.cell(row = ind, column = 5).value = lastDesc
            sheet.cell(row = ind, column = 9).value = lastNotes 

    wb.save('dataframe3.xlsx')
    wb.close()

    """
    ============================================================================
    This portion of code is reponsible for writing the data to databse
    ============================================================================
    """

    # remove previous data
    # =====================
    del_query = ("DELETE FROM trans")
    cursor = CONNECTION.cursor()
    cursor.execute(del_query)
    CONNECTION.commit()   
    
    df = pd.read_excel("dataframe3.xlsx")

    # iterate trhough data rows and assign data values to vaariables
    #  ============================================================== 
    for index, row in islice(df.iterrows(), 1, (beginLast_row_index)): 
        internal_no     =  str(row['InternalNo'])
        date            =  str(row['Date'])
        account         =  str(row['Account'])
        num             =  str(row['Num'])
        description     =  str(row['Description'])
        memo            =  str(row['Memo'])
        category        =  str(row['Category'])
        tag             =  str(row['Tag'])
        notes           =  str(row['Notes'])
        clr             =  str(row['Clr'])
        amount          =  str(row['Amount'])

        # replace nan values with empty data
        # ===================================
        if num == 'nan':
           num = " "
        if memo == 'nan':
           memo = " "
        if tag == 'nan':
           tag = " "
        if notes == 'nan':
           notes = " "


        # define query and data, execute query (write to database)
        # =========================================================
        data = (internal_no, date, account, num, description, memo, category,
               tag, notes, clr, amount)
        save_query = ("INSERT INTO trans (InternalNo, date, account, num, \
                     description, memo, category, tag, notes, clr, amount)"
                     "VALUES (%s, %s, %s, %s, %s, %s, %s, %s,%s, %s, %s)")
        cursor = CONNECTION.cursor()
        cursor.execute(save_query, data)
        CONNECTION.commit()    

    cursor.close()


def main():
    """
    ============================================================================
    Function:       main()
    Purpose:        entry point to program
    Parameter(s):   -None-
    Return:         -None-
    ============================================================================
    """
    global CONNECTION
    preProcess2()
    print("Closing Database Connection . . .")
    CONNECTION.close()
    print("bye . . .")

if __name__ == "__main__":
    main()